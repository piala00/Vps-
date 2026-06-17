"""
NEXORA v2.0 — Gestionnaire de source de donnees
Respecte le choix SQL vs Excel configure dans Parametres.
Applique le cache 24h dans les deux cas.
"""
import os, json, logging
from datetime import datetime, date
from core.database import get_config, db_all

log = logging.getLogger('NEXORA.DataSource')

_CACHE_FILE  = os.path.join(os.path.dirname(__file__), '..', 'Data', 'sage_cache.json')
_CACHE_MAX_H = 24

os.makedirs(os.path.dirname(_CACHE_FILE), exist_ok=True)


def _save_cache(key, data):
    try:
        cache = {}
        if os.path.exists(_CACHE_FILE):
            try:
                with open(_CACHE_FILE,'r',encoding='utf-8') as f:
                    cache = json.load(f)
            except Exception:
                cache = {}
        cache[key] = {'ts': datetime.now().isoformat(), 'data': data}
        with open(_CACHE_FILE,'w',encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, default=str)
    except Exception as e:
        log.warning("Cache save: %s", e)


def _load_cache(key):
    try:
        if not os.path.exists(_CACHE_FILE): return None
        for enc in ('utf-8','cp1252','latin-1'):
            try:
                with open(_CACHE_FILE,'r',encoding=enc) as f:
                    cache = json.load(f)
                break
            except (UnicodeDecodeError, json.JSONDecodeError):
                continue
        else:
            return None
        entry = cache.get(key)
        if not entry: return None
        ts    = datetime.fromisoformat(entry['ts'])
        age_h = (datetime.now()-ts).total_seconds()/3600
        if age_h > _CACHE_MAX_H:
            log.info("Cache %s perime (%.1fh)", key, age_h)
            return None
        log.info("Cache %s charge (%.1fh)", key, age_h)
        return entry['data']
    except Exception as e:
        log.warning("Cache load: %s", e)
        return None


def clear_cache():
    try:
        if os.path.exists(_CACHE_FILE):
            os.remove(_CACHE_FILE)
    except Exception:
        pass


def get_ref_clients():
    """Charge le referentiel clients depuis la DB locale."""
    rows = db_all(
        "SELECT code,nom,zone,commercial,plafond,delai,telephone FROM ref_clients")
    return {r['code']: {
        'nom':        r['nom']       or '',
        'zone':       r['zone']      or '',
        'commercial': r['commercial'] or '',
        'plafond':    r['plafond']   or 0,
        'delai':      r['delai']     or 30,
        'telephone':  r['telephone'] or '',
    } for r in rows}


def get_commerciaux_list():
    """Charge la liste officielle des commerciaux."""
    rows = db_all("SELECT nom,objectif FROM liste_commerciaux ORDER BY ordre,nom")
    return [(r['nom'], r['objectif']) for r in rows]


def load_grand_livre_data(force=False):
    """
    Charge les donnees Grand Livre selon la source configuree.
    - source='sql'   → Sage SQL Server (cache 24h)
    - source='excel' → Fichier Excel configure
    Retourne (all_rows, ref, commerciaux) ou ([], {}, []) si erreur.
    """
    source   = get_config('source', 'sql')
    ref      = get_ref_clients()
    coms     = get_commerciaux_list()

    if source == 'excel':
        path = get_config('excel_path','')
        if not path or not os.path.exists(path):
            log.warning("Excel non configure ou introuvable: %s", path)
            return [], ref, coms
        cache_key = 'excel_' + path
        if not force:
            cached = _load_cache(cache_key)
            if cached is not None:
                log.info("Excel depuis cache: %d lignes", len(cached))
                return _restore_dates(cached), ref, coms
        try:
            from openpyxl import load_workbook
            wb     = load_workbook(path, read_only=True, data_only=True)
            import_rows = []
            for name in wb.sheetnames:
                from unicodedata import normalize, category as ucat
                def _n(s): return ' '.join(''.join(c for c in normalize('NFD',s.upper()) if ucat(c)!='Mn').split())
                if _n(name) == 'IMPORT':
                    for row in wb[name].iter_rows(values_only=True):
                        import_rows.append(tuple(row))
                    break
            wb.close()
            if not import_rows:
                log.warning("Feuille IMPORT non trouvee dans %s", path)
                return [], ref, coms
            from core.commercial_engine import build_raw_from_excel
            rows = build_raw_from_excel(import_rows, ref)
            _save_cache(cache_key, _serialize_rows(rows))
            log.info("Excel charge: %d lignes", len(rows))
            return rows, ref, coms
        except Exception as e:
            log.error("Erreur lecture Excel: %s", e)
            return [], ref, coms

    else:  # source == 'sql' (défaut)
        cache_key = 'sql_grand_livre'
        if not force:
            cached = _load_cache(cache_key)
            if cached is not None:
                log.info("SQL depuis cache: %d lignes", len(cached))
                return _restore_dates(cached), ref, coms
        try:
            from core.sage_connector import sage_query, SQL_GRAND_LIVRE
            sql_rows = sage_query(SQL_GRAND_LIVRE)
            if not sql_rows:
                log.warning("Aucune donnee SQL depuis Sage")
                return [], ref, coms
            from core.commercial_engine import build_raw_from_sql
            rows = build_raw_from_sql(sql_rows, ref)
            _save_cache(cache_key, _serialize_rows(rows))
            log.info("SQL charge: %d lignes", len(rows))
            return rows, ref, coms
        except Exception as e:
            log.error("Erreur SQL Sage: %s", e)
            return [], ref, coms


def _serialize_rows(rows):
    """Serialise les dates pour JSON."""
    out = []
    for r in rows:
        row = dict(r)
        if isinstance(row.get('date'), date):
            row['date'] = row['date'].isoformat()
        out.append(row)
    return out


def _restore_dates(rows):
    """Restaure les dates depuis JSON."""
    for r in rows:
        d = r.get('date')
        if isinstance(d, str) and d:
            try:    r['date'] = date.fromisoformat(d)
            except: r['date'] = None
    return rows
