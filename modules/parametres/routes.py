"""
NEXORA v2.0 — Module Parametres (COMPLET)
Tout ce qui existe dans GTC ERP PILOT V3 AdminFrame est ici :
- Connexion Sage robuste (SQL/Excel/Windows auth)
- Base de donnees applicative (SQLite local/reseau/SQL Server)
- Utilisateurs avec poste/role/commercial_name/telegram_id + PBKDF2
- Referentiel clients avec import Excel
- Commerciaux
- Bot Telegram + demarrage
- Journal de processus
- Compilation .exe / PyInstaller
- Tous les modules heritent de la config via get_config()
"""
import os, socket, logging, subprocess, sys
from flask import render_template, jsonify, session, request
from . import bp
from core.auth import login_required, get_user
from core.database import (get_config, set_config, PERMISSIONS_TREE,
                            db_all, db_one, db_exec, hash_pwd, verify_pwd,
                            audit, get_accessible_modules, set_all_permissions)

log = logging.getLogger('NEXORA.Parametres')


# ── Page principale ───────────────────────────────────────────────────────────

@bp.route('/module/parametres')
@login_required
def module_parametres():
    user    = get_user()
    modules = session.get('modules', [])
    cfg     = {'nom_societe': get_config('nom_societe', ''),
               'devise':      get_config('devise', 'XAF')}
    return render_template('modules/parametres.html',
        module='parametres', module_label='Parametres',
        user=user, modules=modules, config=cfg,
        tree=PERMISSIONS_TREE)


# ── Configuration generale (source SQL/Excel + enregistrer) ───────────────────

@bp.route('/api/config/general', methods=['GET', 'POST'])
@login_required
def api_config_general():
    if request.method == 'POST':
        d = request.get_json() or {}
        keys = ['source', 'excel_path', 'sage_server', 'sage_database',
                'sage_trusted', 'sage_user', 'sage_password', 'sage_annee',
                'nom_societe', 'devise', 'ville_siege', 'telephone']
        for k in keys:
            if k in d:
                set_config(k, str(d[k]))
        audit(session.get('username','?'), 'PARAM', 'CONFIG_SAVE', '')
        return jsonify({'ok': True, 'message': 'Configuration enregistree'})
    return jsonify({'ok': True, 'config': {
        'source':        get_config('source', 'sql'),
        'excel_path':    get_config('excel_path', ''),
        'sage_server':   get_config('sage_server', ''),
        'sage_database': get_config('sage_database', ''),
        'sage_trusted':  get_config('sage_trusted', '1'),
        'sage_user':     get_config('sage_user', 'sa'),
        'sage_annee':    get_config('sage_annee', '2026'),
        'nom_societe':   get_config('nom_societe', ''),
        'devise':        get_config('devise', 'XAF'),
    }})


# ── Test connexion Sage robuste ───────────────────────────────────────────────

@bp.route('/api/config/sage', methods=['GET', 'POST'])
@login_required
def api_config_sage():
    if request.method == 'POST':
        d = request.get_json() or {}
        for k in ['sage_server','sage_database','sage_user','sage_password','sage_trusted','sage_annee']:
            if k in d:
                set_config(k, str(d[k]))
        return jsonify({'ok': True})
    return jsonify({'ok': True, 'config': {
        'sage_server':   get_config('sage_server', ''),
        'sage_database': get_config('sage_database', ''),
        'sage_user':     get_config('sage_user', 'sa'),
        'sage_trusted':  get_config('sage_trusted', '1'),
        'sage_annee':    get_config('sage_annee', '2026'),
    }})


@bp.route('/api/config/test-sage-robuste', methods=['POST'])
@login_required
def api_test_sage_robuste():
    """Test de connexion SQL Server avec diagnostic complet (tous drivers + TCP)."""
    d       = request.get_json() or {}
    server  = d.get('sage_server','').strip()   or get_config('sage_server','')
    db_name = d.get('sage_database','').strip() or get_config('sage_database','')
    trusted = d.get('sage_trusted', True)
    user    = d.get('sage_user','sa')           or get_config('sage_user','sa')
    pw      = d.get('sage_password','')         or get_config('sage_password','')
    from core.sage_connector import test_connection
    ok, msg = test_connection(server, db_name, trusted, user, pw)
    if ok:
        set_config('sage_server',   server)
        set_config('sage_database', db_name)
        set_config('sage_trusted',  '1' if trusted else '0')
        set_config('sage_user',     user)
        set_config('sage_password', pw)
        audit(session.get('username','?'), 'PARAM', 'SAGE_TEST_OK', server)
    return jsonify({'ok': ok, 'message': msg})


# ── Import depuis Excel ───────────────────────────────────────────────────────

@bp.route('/api/config/upload-excel', methods=['POST'])
@login_required
def api_upload_excel():
    """
    Recoit un fichier .xlsx envoye depuis le navigateur (multipart/form-data)
    et le sauvegarde sur le serveur. Remplace l'ancienne logique desktop qui
    demandait de taper un chemin local (incompatible avec une app web).
    """
    if 'file' not in request.files:
        return jsonify({'ok': False, 'message': 'Aucun fichier recu'})
    f = request.files['file']
    if not f or not f.filename:
        return jsonify({'ok': False, 'message': 'Nom de fichier vide'})
    if not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'ok': False, 'message': 'Seuls les fichiers .xlsx/.xlsm sont acceptes'})

    upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'Data', 'uploads')
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = 'source_excel.xlsx'
    dest_path = os.path.normpath(os.path.join(upload_dir, safe_name))
    f.save(dest_path)

    set_config('excel_path', dest_path)
    audit(session.get('username', '?'), 'PARAM', 'EXCEL_UPLOAD', f.filename)
    return jsonify({'ok': True, 'message': f'Fichier "{f.filename}" recu et enregistre.',
                    'path': dest_path})


def _norm_sheet(s):
    """Normalise un nom de feuille/colonne : majuscules, sans accents, espaces uniformises."""
    from unicodedata import normalize, category as ucat
    s = (s or '').upper()
    s = ''.join(c for c in normalize('NFD', s) if ucat(c) != 'Mn')
    s = s.replace('_', ' ')
    return ' '.join(s.split())


@bp.route('/api/config/import-excel', methods=['POST'])
@login_required
def api_import_excel():
    """
    Import depuis le fichier Excel deja uploade (ou chemin fourni en fallback).
    Lit les feuilles REF_CLIENTS et LISTE_COMMERCIEAUX (variantes de nom tolerees),
    selon le format reel des fichiers GTC.
    """
    d    = request.get_json() or {}
    path = d.get('excel_path', '').strip() or get_config('excel_path', '')
    if not path or not os.path.exists(path):
        return jsonify({'ok': False, 'message': 'Aucun fichier Excel disponible. Importez-le d\'abord via le bouton ci-dessus.'})
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        nb_ref = 0; nb_com = 0

        # ── Feuille REF_CLIENTS (tolere les variantes de nom) ──
        ref_sheet = next((s for s in wb.sheetnames if _norm_sheet(s) == 'REF CLIENTS'), None)
        if ref_sheet:
            ws = wb[ref_sheet]
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]:
                    continue
                code = str(row[0]).strip()
                if not code or _norm_sheet(code) in ('CODE', 'CODE CLIENT'):
                    continue
                db_exec(
                    "INSERT OR REPLACE INTO ref_clients(code,nom,zone,commercial,plafond,delai,telephone)"
                    " VALUES(?,?,?,?,?,?,?)",
                    (code,
                     str(row[1]).strip() if len(row) > 1 and row[1] else '',
                     str(row[2]).strip() if len(row) > 2 and row[2] else '',
                     str(row[3]).strip() if len(row) > 3 and row[3] else '',
                     float(row[4]) if len(row) > 4 and row[4] else 0,
                     int(float(row[5])) if len(row) > 5 and row[5] else 30,
                     str(row[6]).strip() if len(row) > 6 and row[6] else ''))
                nb_ref += 1
        else:
            log.warning("Feuille REF_CLIENTS introuvable dans %s (feuilles: %s)", path, wb.sheetnames)

        # ── Feuille LISTE_COMMERCIEAUX (tolere variantes orthographiques) ──
        EXCLUS = {'COMMERCIAL', 'TOTAL', 'OBJECTIF', 'NOM', 'AUTRE CLIENT',
                  'LISTE COMMERCIEAUX', 'LISTE COMMERCIAUX', ''}
        com_sheet = next(
            (s for s in wb.sheetnames
             if _norm_sheet(s) in ('LISTE COMMERCIEAUX', 'LISTE COMMERCIAUX', 'COMMERCIAUX')),
            None)
        if com_sheet:
            ws = wb[com_sheet]
            for row in ws.iter_rows(values_only=True):
                if not row or len(row) < 2:
                    continue
                nom_brut = row[1]
                if nom_brut is None:
                    continue
                nom = str(nom_brut).strip()
                if not nom or _norm_sheet(nom) in EXCLUS:
                    continue
                # La colonne objectif est en position 2 (3eme colonne)
                obj = float(row[2]) if len(row) > 2 and isinstance(row[2], (int, float)) else 0
                db_exec(
                    "INSERT OR REPLACE INTO liste_commerciaux(nom,objectif,ordre) VALUES(?,?,?)",
                    (nom, obj, nb_com))
                nb_com += 1
        else:
            log.warning("Feuille LISTE_COMMERCIEAUX introuvable dans %s (feuilles: %s)", path, wb.sheetnames)

        wb.close()
        set_config('excel_path', path)
        set_config('source', 'excel')
        audit(session.get('username', '?'), 'PARAM', 'EXCEL_IMPORT', f'ref={nb_ref} com={nb_com}')

        if nb_ref == 0 and nb_com == 0:
            return jsonify({'ok': False,
                            'message': f'Aucune donnee importee. Feuilles trouvees dans le fichier : {", ".join(wb.sheetnames)}'})
        return jsonify({'ok': True,
                        'message': f'{nb_ref} clients et {nb_com} commerciaux importes',
                        'nb_ref': nb_ref, 'nb_com': nb_com})
    except Exception as e:
        log.error("Import Excel: %s", e)
        return jsonify({'ok': False, 'message': str(e)})


# ── Base de donnees applicative ───────────────────────────────────────────────

@bp.route('/api/database/sqlite-info')
@login_required
def api_sqlite_info():
    from config import Config
    path = Config.DB_PATH
    try:
        size      = os.path.getsize(path)
        size_h    = (str(round(size/1024,0)) + ' Ko') if size < 1024*1024 else (str(round(size/1024/1024,1)) + ' Mo')
        nb_tables = db_one("SELECT COUNT(*) n FROM sqlite_master WHERE type='table'") or {'n':0}
        return jsonify({'ok':True,'path':path,'size':size,'size_human':size_h,
                        'nb_tables':nb_tables['n']})
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)})


@bp.route('/api/database/backup-sqlite', methods=['POST'])
@login_required
def api_backup_sqlite():
    import shutil
    from config import Config
    from datetime import datetime
    try:
        ts  = datetime.now().strftime('%Y%m%d_%H%M%S')
        dst = Config.DB_PATH.replace('.db', '_backup_' + ts + '.db')
        shutil.copy2(Config.DB_PATH, dst)
        audit(session.get('username','?'), 'DB', 'BACKUP', dst)
        return jsonify({'ok':True,'path':dst,'message':'Sauvegarde creee'})
    except Exception as e:
        return jsonify({'ok':False,'message':str(e)})


@bp.route('/api/database/create-app-db', methods=['POST'])
@login_required
def api_create_app_db():
    """
    Crée / initialise la base applicative.
    SQLite local, SQLite réseau (chemin UNC) ou SQL Server.
    """
    d       = request.get_json() or {}
    db_type = d.get('db_type', 'SQLite local')
    path    = d.get('db_path', '').strip()
    db_name = d.get('db_name', 'NEXORA_APP').strip() or 'NEXORA_APP'
    try:
        if 'SQL Server' in db_type:
            try:
                import pyodbc
                srv = path or get_config('sage_server', '')
                cs  = (f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={srv};'
                       f'DATABASE=master;Trusted_Connection=yes;')
                cx  = pyodbc.connect(cs, timeout=10, autocommit=True)
                cx.cursor().execute(
                    f"IF NOT EXISTS(SELECT name FROM sys.databases WHERE name=N'{db_name}')"
                    f" CREATE DATABASE [{db_name}]")
                cx.close()
                set_config('app_db_type',   'sqlserver')
                set_config('app_db_server',  srv)
                set_config('app_db_name',    db_name)
                audit(session.get('username','?'), 'DB', 'CREATE_SQLSERVER', db_name)
                return jsonify({'ok':True,'message': f"Base SQL Server '{db_name}' creee sur {srv}"})
            except Exception as e:
                return jsonify({'ok':False,'message':str(e)})
        else:
            import sqlite3 as _s3
            db_path = path if path else os.path.join(
                os.path.dirname(os.path.abspath(__file__)), '..', '..', 'nexora_app.db')
            os.makedirs(os.path.dirname(db_path) if os.path.dirname(db_path) else '.', exist_ok=True)
            cx = _s3.connect(db_path, timeout=10)
            cx.execute("PRAGMA journal_mode=WAL")
            cx.execute("CREATE TABLE IF NOT EXISTS nexora_app_users(id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE NOT NULL, role TEXT DEFAULT 'commercial')")
            cx.execute("CREATE TABLE IF NOT EXISTS nexora_app_config(key TEXT PRIMARY KEY, value TEXT DEFAULT '')")
            cx.commit(); cx.close()
            set_config('app_db_type', 'sqlite')
            set_config('app_db_path', db_path)
            audit(session.get('username','?'), 'DB', 'CREATE_SQLITE', db_path)
            return jsonify({'ok':True,'message':f"Base SQLite creee : {db_path}"})
    except Exception as e:
        return jsonify({'ok':False,'message':str(e)})


@bp.route('/api/database/test-app-db', methods=['POST'])
@login_required
def api_test_app_db():
    d       = request.get_json() or {}
    db_type = d.get('db_type', 'SQLite local')
    path    = d.get('db_path', '').strip()
    db_name = d.get('db_name', 'NEXORA_APP').strip()
    try:
        if 'SQL Server' in db_type:
            import pyodbc
            srv = path or get_config('sage_server','')
            cs  = (f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={srv};'
                   f'DATABASE={db_name or "master"};Trusted_Connection=yes;')
            cx  = pyodbc.connect(cs, timeout=5)
            cx.close()
            return jsonify({'ok':True,'message': f'SQL Server OK : {srv}/{db_name}'})
        else:
            import sqlite3 as _s3
            db_path = path or 'nexora_app.db'
            cx = _s3.connect(db_path, timeout=5)
            cx.execute("SELECT 1"); cx.close()
            return jsonify({'ok':True,'message':f'SQLite OK : {db_path}'})
    except Exception as e:
        return jsonify({'ok':False,'message':str(e)})


# ── Utilisateurs (complets avec poste/role/commercial_name/telegram_id) ───────

@bp.route('/api/utilisateurs', methods=['GET', 'POST'])
@login_required
def api_utilisateurs():
    if request.method == 'POST':
        d        = request.get_json() or {}
        nom      = d.get('nom','').strip()
        username = d.get('username','').strip()
        pwd      = d.get('password','')
        if not nom or not username:
            return jsonify({'ok':False,'msg':'Nom et login obligatoires'})
        existing = db_one("SELECT id FROM utilisateurs WHERE username=?", (username,))
        if existing:
            return jsonify({'ok':False,'msg':'Login deja utilise'})
        uid = db_exec(
            "INSERT INTO utilisateurs(username,password_hash,nom,prenom,agence,role,"
            "poste,categorie,commercial_name,telegram_id,actif) VALUES(?,?,?,?,?,?,?,?,?,?,1)",
            (username, hash_pwd(pwd) if pwd else '',
             nom, d.get('prenom',''), d.get('agence','BERTOUA'),
             d.get('role','commercial'), d.get('poste',''),
             d.get('categorie',''), d.get('commercial_name',''),
             d.get('telegram_id','')))
        set_all_permissions(uid, True)
        audit(session.get('username','?'), 'USER', 'CREATE', username)
        return jsonify({'ok':True,'id':uid})
    users = db_all(
        "SELECT id,username,nom,prenom,agence,role,poste,categorie,"
        "commercial_name,telegram_id,actif FROM utilisateurs ORDER BY nom")
    return jsonify({'ok':True,'utilisateurs':users})


@bp.route('/api/utilisateurs/<int:uid>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_utilisateur(uid):
    if request.method == 'GET':
        u = db_one("SELECT * FROM utilisateurs WHERE id=?", (uid,))
        return jsonify({'ok':True,'utilisateur':dict(u) if u else {}})
    if request.method == 'DELETE':
        if uid == 1:
            return jsonify({'ok':False,'msg':'Impossible de supprimer l\'administrateur'})
        db_exec("DELETE FROM utilisateurs WHERE id=?", (uid,))
        return jsonify({'ok':True})
    # PUT — Modifier
    d = request.get_json() or {}
    pwd = d.get('password','')
    if pwd:
        db_exec(
            "UPDATE utilisateurs SET nom=?,prenom=?,agence=?,role=?,poste=?,"
            "categorie=?,commercial_name=?,telegram_id=?,actif=?,password_hash=? WHERE id=?",
            (d.get('nom',''), d.get('prenom',''), d.get('agence','BERTOUA'),
             d.get('role','commercial'), d.get('poste',''), d.get('categorie',''),
             d.get('commercial_name',''), d.get('telegram_id',''),
             d.get('actif',1), hash_pwd(pwd), uid))
    else:
        db_exec(
            "UPDATE utilisateurs SET nom=?,prenom=?,agence=?,role=?,poste=?,"
            "categorie=?,commercial_name=?,telegram_id=?,actif=? WHERE id=?",
            (d.get('nom',''), d.get('prenom',''), d.get('agence','BERTOUA'),
             d.get('role','commercial'), d.get('poste',''), d.get('categorie',''),
             d.get('commercial_name',''), d.get('telegram_id',''),
             d.get('actif',1), uid))
    audit(session.get('username','?'), 'USER', 'UPDATE', str(uid))
    return jsonify({'ok':True})


# ── Droits d'accès ────────────────────────────────────────────────────────────

@bp.route('/api/utilisateurs/<int:uid>/droits', methods=['GET', 'POST'])
@login_required
def api_droits_user(uid):
    if request.method == 'POST':
        d      = request.get_json() or {}
        droits = d.get('droits', {})
        for mod, autorise in droits.items():
            db_exec("DELETE FROM permissions WHERE user_id=? AND module=?", (uid, mod))
            if autorise:
                db_exec(
                    "INSERT OR REPLACE INTO permissions(user_id,module,sous_module,action,autorise)"
                    " VALUES(?,?,'*','view',1)", (uid, mod))
        audit(session.get('username','?'), 'DROITS', 'UPDATE', 'user='+str(uid))
        return jsonify({'ok':True,'message':'Droits mis a jour'})
    mods_actifs = [r['module'] for r in
        db_all("SELECT DISTINCT module FROM permissions WHERE user_id=? AND autorise=1", (uid,))]
    mods = {m:(m in mods_actifs) for m in PERMISSIONS_TREE}
    return jsonify({'ok':True,'modules':mods})


@bp.route('/api/utilisateurs/<int:uid>/permissions/tout', methods=['POST'])
@login_required
def api_permissions_tout(uid):
    d   = request.get_json() or {}
    val = d.get('tout', True)
    set_all_permissions(uid, val)
    return jsonify({'ok':True})


# ── Droits d'accès detailles (module -> sous-module -> lecture/ecriture/suppression) ──

@bp.route('/api/utilisateurs/<int:uid>/permissions-detail')
@login_required
def api_permissions_detail(uid):
    """
    Retourne l'arborescence complete des permissions fines de l'utilisateur,
    structuree par module puis sous-module avec le detail
    lecture/ecriture/suppression/tout.
    """
    from core.database import get_user_permissions_full, PERMISSION_ACTIONS
    perms = get_user_permissions_full(uid)
    return jsonify({'ok':True,'permissions':perms,'actions':PERMISSION_ACTIONS,
                    'tree':PERMISSIONS_TREE})


@bp.route('/api/utilisateurs/<int:uid>/permissions-detail', methods=['POST'])
@login_required
def api_permissions_detail_save(uid):
    """
    Sauvegarde les permissions fines.
    Payload attendu : {
        "permissions": {
            "commercial": {"com-cockpit": {"lecture":true,"ecriture":false,"suppression":false}, ...},
            "stock": {"*": {"lecture":true, ...}},
            ...
        }
    }
    """
    from core.database import set_permission_detail
    d     = request.get_json() or {}
    perms = d.get('permissions', {})
    for mod, sous_mods in perms.items():
        for sm, actions in sous_mods.items():
            set_permission_detail(uid, mod, sm, actions)
    audit(session.get('username','?'), 'DROITS', 'UPDATE_DETAIL', 'user='+str(uid))
    return jsonify({'ok':True,'message':'Permissions detaillees mises a jour'})


# ── Référentiel Clients ───────────────────────────────────────────────────────

@bp.route('/api/param/ref-clients', methods=['GET', 'POST', 'DELETE'])
@login_required
def api_ref_clients():
    if request.method == 'POST':
        d    = request.get_json() or {}
        code = d.get('code','').strip()
        if not code:
            return jsonify({'ok':False,'msg':'Code client obligatoire'})
        db_exec(
            "INSERT OR REPLACE INTO ref_clients(code,nom,zone,commercial,plafond,delai,telephone)"
            " VALUES(?,?,?,?,?,?,?)",
            (code, d.get('nom',''), d.get('zone',''), d.get('commercial',''),
             float(d.get('plafond',0)), int(d.get('delai',30)),
             d.get('telephone','')))
        return jsonify({'ok':True})
    if request.method == 'DELETE':
        code = request.args.get('code','')
        if code:
            db_exec("DELETE FROM ref_clients WHERE code=?", (code,))
        return jsonify({'ok':True})
    q   = request.args.get('q','')
    sql = "SELECT * FROM ref_clients"
    p   = []
    if q:
        sql += " WHERE nom LIKE ? OR code LIKE ? OR commercial LIKE ?"
        p    = ['%'+q+'%','%'+q+'%','%'+q+'%']
    sql += " ORDER BY code"
    return jsonify({'ok':True,'clients':db_all(sql, p)})


@bp.route('/api/param/ref-clients/import', methods=['POST'])
@login_required
def api_ref_clients_import():
    """Import en masse depuis JSON (liste de lignes)."""
    d    = request.get_json() or {}
    rows = d.get('rows', [])
    nb   = 0
    for r in rows:
        code = str(r[0]).strip() if r else ''
        if not code: continue
        db_exec(
            "INSERT OR REPLACE INTO ref_clients(code,nom,zone,commercial,plafond,delai,telephone)"
            " VALUES(?,?,?,?,?,?,?)",
            (code, str(r[1] if len(r)>1 else ''),
             str(r[2] if len(r)>2 else ''),
             str(r[3] if len(r)>3 else ''),
             float(r[4]) if len(r)>4 and r[4] else 0,
             int(r[5])   if len(r)>5 and r[5] else 30,
             str(r[6] if len(r)>6 else '')))
        nb += 1
    audit(session.get('username','?'), 'PARAM', 'REF_IMPORT', f'nb={nb}')
    return jsonify({'ok':True,'nb_importes':nb})


# ── Commerciaux ───────────────────────────────────────────────────────────────

@bp.route('/api/param/commerciaux', methods=['GET', 'POST', 'DELETE'])
@login_required
def api_commerciaux():
    if request.method == 'POST':
        d   = request.get_json() or {}
        nom = d.get('nom','').strip()
        if not nom:
            return jsonify({'ok':False,'msg':'Nom obligatoire'})
        db_exec(
            "INSERT OR REPLACE INTO liste_commerciaux(nom,objectif,ordre) VALUES(?,?,?)",
            (nom, float(d.get('objectif',0)), int(d.get('ordre',0))))
        return jsonify({'ok':True})
    if request.method == 'DELETE':
        nom = request.args.get('nom','')
        if nom:
            db_exec("DELETE FROM liste_commerciaux WHERE nom=?", (nom,))
        return jsonify({'ok':True})
    rows = db_all("SELECT nom,objectif,ordre FROM liste_commerciaux ORDER BY ordre,nom")
    return jsonify({'ok':True,'commerciaux':rows})


# ── Bot Telegram ──────────────────────────────────────────────────────────────

@bp.route('/api/param/bot-telegram', methods=['GET', 'POST'])
@login_required
def api_bot_telegram():
    if request.method == 'POST':
        d = request.get_json() or {}
        for k, v in d.items():
            db_exec("INSERT OR REPLACE INTO bot_config(key,value) VALUES(?,?)", (k, str(v)))
        audit(session.get('username','?'), 'PARAM', 'BOT_CONFIG', '')
        return jsonify({'ok':True})
    rows = db_all("SELECT key, value FROM bot_config")
    return jsonify({'ok':True,'config':{r['key']:r['value'] for r in rows}})


@bp.route('/api/param/bot-telegram/test', methods=['POST'])
@login_required
def api_bot_test():
    rows  = db_all("SELECT key, value FROM bot_config")
    cfg   = {r['key']:r['value'] for r in rows}
    token = cfg.get('bot_token','')
    admins= cfg.get('bot_admin_ids','')
    if not token:
        return jsonify({'ok':False,'msg':'Token Telegram non configure'})
    try:
        import urllib.request, urllib.parse, json as _json
        sent = 0
        for admin_id in [x.strip() for x in admins.split(',') if x.strip()]:
            url  = f'https://api.telegram.org/bot{token}/sendMessage'
            data = urllib.parse.urlencode({
                'chat_id': admin_id,
                'text':    'NEXORA v2.0 — Test connexion Bot Telegram OK ✅',
            }).encode()
            req  = urllib.request.Request(url, data=data)
            resp = urllib.request.urlopen(req, timeout=10)
            res  = _json.loads(resp.read())
            if res.get('ok'):
                sent += 1
        return jsonify({'ok':True,'message': f'Message envoye a {sent} admin(s)'})
    except Exception as e:
        return jsonify({'ok':False,'msg':str(e)})


@bp.route('/api/param/bot-telegram/demarrer', methods=['POST'])
@login_required
def api_bot_demarrer():
    """Demarre le vrai bot Telegram (BotManager) en thread daemon."""
    rows  = db_all("SELECT key, value FROM bot_config")
    cfg   = {r['key']:r['value'] for r in rows}
    token = cfg.get('bot_token','')
    if not token:
        return jsonify({'ok':False,'msg':'Token Telegram non configure. Sauvegardez d\'abord.'})
    try:
        from core.telegram_bot import get_bot_manager
        mgr = get_bot_manager()
        if mgr.running:
            return jsonify({'ok':True,'message':'Bot deja actif.'})
        mgr.start(token)
        audit(session.get('username','?'), 'BOT', 'START', '')
        return jsonify({'ok':True,'message':'Demarrage du bot lance. Verifiez le statut dans quelques secondes.'})
    except Exception as e:
        return jsonify({'ok':False,'msg':str(e)})


@bp.route('/api/param/bot-telegram/arreter', methods=['POST'])
@login_required
def api_bot_arreter():
    try:
        from core.telegram_bot import get_bot_manager
        mgr = get_bot_manager()
        mgr.stop()
        audit(session.get('username','?'), 'BOT', 'STOP', '')
        return jsonify({'ok':True,'message':'Bot arrete.'})
    except Exception as e:
        return jsonify({'ok':False,'msg':str(e)})


@bp.route('/api/param/bot-telegram/statut')
@login_required
def api_bot_statut():
    try:
        from core.telegram_bot import get_bot_manager, get_bot_log
        mgr = get_bot_manager()
        return jsonify({'ok':True,'running':mgr.running,'log':get_bot_log(30)})
    except Exception as e:
        return jsonify({'ok':False,'msg':str(e)})


# ── Inscriptions Bot Telegram ──────────────────────────────────────────────────

@bp.route('/api/param/bot-inscriptions')
@login_required
def api_bot_inscriptions():
    rows = db_all(
        "SELECT id,telegram_id,telegram_nom,nom,poste,agence,statut,created_at "
        "FROM bot_inscriptions ORDER BY id DESC LIMIT 100")
    return jsonify({'ok':True,'inscriptions':rows})


@bp.route('/api/param/bot-inscriptions/<int:iid>/valider', methods=['POST'])
@login_required
def api_bot_inscription_valider(iid):
    """Valide une inscription : cree l'utilisateur NEXORA et lie le Telegram ID."""
    d    = request.get_json() or {}
    insc = db_one("SELECT * FROM bot_inscriptions WHERE id=?", (iid,))
    if not insc:
        return jsonify({'ok':False,'msg':'Inscription introuvable'})
    username = d.get('username','').strip() or (insc['telegram_id'])
    role     = d.get('role','commercial')
    existing = db_one("SELECT id FROM utilisateurs WHERE username=?", (username,))
    if not existing:
        db_exec(
            "INSERT INTO utilisateurs(username,password_hash,nom,prenom,agence,role,"
            "poste,categorie,commercial_name,telegram_id,actif) VALUES(?,?,?,?,?,?,?,?,?,?,1)",
            (username, '', insc['nom'], '', insc['agence'], role,
             insc['poste'], '', d.get('commercial_name',''), insc['telegram_id']))
    else:
        db_exec("UPDATE utilisateurs SET telegram_id=? WHERE id=?",
                (insc['telegram_id'], existing['id']))
    db_exec(
        "UPDATE bot_inscriptions SET statut='VALIDEE',validated_by=?,validated_at=datetime('now') WHERE id=?",
        (session.get('username','?'), iid))
    audit(session.get('username','?'), 'BOT', 'INSCRIPTION_VALIDEE', username)
    return jsonify({'ok':True,'message':'Utilisateur cree/mis a jour et inscription validee'})


@bp.route('/api/param/bot-inscriptions/<int:iid>/rejeter', methods=['POST'])
@login_required
def api_bot_inscription_rejeter(iid):
    db_exec("UPDATE bot_inscriptions SET statut='REJETEE',validated_by=?,validated_at=datetime('now') WHERE id=?",
            (session.get('username','?'), iid))
    return jsonify({'ok':True})


# ── Snapshot automatique (pour que le bot lise les KPIs/Classement/Creances) ──

@bp.route('/api/param/bot-telegram/snapshot-auto', methods=['POST'])
@login_required
def api_bot_snapshot_auto():
    """
    Calcule kpis/classement/creances/alertes via le moteur commercial
    et les sauvegarde dans data_snapshot pour que le bot Telegram les lise.
    """
    import json as _json
    from core.data_source import load_grand_livre_data
    from core.commercial_engine import init_commerciaux, compute_all
    all_rows, ref, coms = load_grand_livre_data(force=False)
    if not all_rows:
        return jsonify({'ok':False,'msg':'Aucune donnee disponible (Sage/Excel)'})
    init_commerciaux(coms)
    data = compute_all(all_rows, ref)

    def _default(o):
        from datetime import date, datetime as dt
        if isinstance(o,(date,dt)): return str(o)
        try: return float(o)
        except Exception: return str(o)

    snapshots = {
        'kpis':       data['kpis'],
        'classement': data['classement'],
        'creances':   data['creances'],
        'alertes':    data['alertes'],
        'period':     {'loaded_at': data.get('loaded_at',''), 'label': data.get('period_label','')},
    }
    for k, v in snapshots.items():
        db_exec(
            "INSERT OR REPLACE INTO data_snapshot(key,value,updated_at) VALUES(?,?,datetime('now'))",
            (k, _json.dumps(v, ensure_ascii=False, default=_default)))
    audit(session.get('username','?'), 'BOT', 'SNAPSHOT', f"{len(data['classement'])} commerciaux")
    return jsonify({'ok':True,'message':'Snapshot mis a jour pour le bot',
                    'nb_commerciaux':len(data['classement']),
                    'nb_creances':len(data['creances'])})


# ── Journal de processus ──────────────────────────────────────────────────────

@bp.route('/api/param/journal', methods=['GET', 'DELETE'])
@login_required
def api_journal():
    if request.method == 'DELETE':
        db_exec("DELETE FROM audit_log")
        return jsonify({'ok':True})
    limite = int(request.args.get('limite', 100))
    rows   = db_all(
        "SELECT date_op, username, module, action, detail "
        "FROM audit_log ORDER BY id DESC LIMIT ?", (limite,))
    return jsonify({'ok':True,'entries':rows})


# ── Journal Sage temps reel (diagnostic blocage connexion) ───────────────────

@bp.route('/api/param/journal-sage', methods=['GET', 'DELETE'])
@login_required
def api_journal_sage():
    """
    Journal en temps reel de chaque tentative de connexion Sage : driver
    essaye, variante serveur, encodage, succes/echec, duree. Permet de
    voir EXACTEMENT pourquoi un module reste bloque en chargement.
    """
    from core.sage_connector import get_sage_log, clear_sage_log, get_last_good_connection
    if request.method == 'DELETE':
        clear_sage_log()
        return jsonify({'ok': True})
    limite = int(request.args.get('limite', 100))
    return jsonify({
        'ok': True,
        'entries': get_sage_log(limite),
        'derniere_connexion_ok': get_last_good_connection(),
    })


# ── PyInstaller / Compilation ─────────────────────────────────────────────────

@bp.route('/api/param/installer-pyinstaller', methods=['POST'])
@login_required
def api_installer_pyinstaller():
    """Installe PyInstaller via pip."""
    try:
        result = subprocess.run(
            [sys.executable, '-m', 'pip', 'install', 'pyinstaller', '--quiet'],
            capture_output=True, text=True, timeout=120)
        if result.returncode == 0:
            return jsonify({'ok':True,'message':'PyInstaller installe avec succes !'})
        return jsonify({'ok':False,'message':result.stderr[:500] or 'Erreur installation'})
    except Exception as e:
        return jsonify({'ok':False,'message':str(e)})


@bp.route('/api/param/compiler-exe', methods=['POST'])
@login_required
def api_compiler_exe():
    """Lance la compilation NEXORA en .exe via PyInstaller."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    run_py   = os.path.join(base_dir, '..', '..', 'run.py')
    run_py   = os.path.normpath(run_py)
    if not os.path.exists(run_py):
        return jsonify({'ok':False,'message': f'run.py introuvable : {run_py}'})
    dist_dir = os.path.join(os.path.dirname(run_py), 'dist')
    try:
        cmd = [sys.executable, '-m', 'PyInstaller',
               '--onefile', '--windowed',
               '--name', 'NEXORA_v2',
               '--distpath', dist_dir,
               '--noconfirm', '--clean',
               run_py]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=600)
        if result.returncode == 0:
            return jsonify({'ok':True,'message': f'NEXORA.exe compile dans {dist_dir}'})
        err = (result.stderr or result.stdout or 'Erreur inconnue')[-500:]
        return jsonify({'ok':False,'message':err})
    except FileNotFoundError:
        return jsonify({'ok':False,'message':'PyInstaller non installe. Cliquez Installer d\'abord.'})
    except Exception as e:
        return jsonify({'ok':False,'message':str(e)})


# ── Accès internet ────────────────────────────────────────────────────────────

@bp.route('/api/parametres/domaine', methods=['GET', 'POST'])
@login_required
def api_domaine():
    if request.method == 'POST':
        d = request.get_json() or {}
        for k, v in d.items():
            set_config('domaine_'+k, str(v))
        return jsonify({'ok':True})
    return jsonify({'ok':True,'config':{
        'domaine':  get_config('domaine_domaine',''),
        'sous_dom': get_config('domaine_sous_dom',''),
        'mode':     get_config('domaine_mode','local'),
        'port':     get_config('domaine_port','5050'),
    }})


@bp.route('/api/parametres/tester-domaine')
@login_required
def api_tester_domaine():
    domaine = request.args.get('domaine','')
    if not domaine:
        return jsonify({'ok':False,'message':'Domaine vide'})
    try:
        ip = socket.gethostbyname(domaine)
        return jsonify({'ok':True,'ip':ip,'message':'Resolu -> '+ip})
    except Exception as e:
        return jsonify({'ok':False,'message':str(e)})


@bp.route('/api/parametres/vpn/wireguard', methods=['POST'])
@login_required
def api_vpn_wg():
    d = request.get_json() or {}
    set_config('vpn_wireguard_conf', d.get('config',''))
    return jsonify({'ok':True})


@bp.route('/api/parametres/vpn/openvpn', methods=['POST'])
@login_required
def api_vpn_ovpn():
    d = request.get_json() or {}
    set_config('vpn_openvpn_conf', d.get('config',''))
    return jsonify({'ok':True})


@bp.route('/api/config/cloudflare/start', methods=['POST'])
@login_required
def api_cf_start():
    d = request.get_json() or {}
    if d.get('token'):
        set_config('cloudflare_token', d['token'])
    return jsonify({'ok':True,'message':'Cloudflare configure'})


@bp.route('/api/config/cloudflare/stop', methods=['POST'])
@login_required
def api_cf_stop():
    return jsonify({'ok':True})


@bp.route('/api/config/ngrok/start', methods=['POST'])
@login_required
def api_ngrok_start():
    d = request.get_json() or {}
    if d.get('token'):
        set_config('ngrok_token', d['token'])
    return jsonify({'ok':True,'message':'ngrok configure'})


@bp.route('/api/config/ngrok/stop', methods=['POST'])
@login_required
def api_ngrok_stop():
    return jsonify({'ok':True})


# ── IP Whitelist ──────────────────────────────────────────────────────────────

@bp.route('/api/config/ip-whitelist', methods=['GET', 'POST'])
@login_required
def api_ip_whitelist():
    if request.method == 'POST':
        d   = request.get_json() or {}
        ips = get_config('ip_whitelist','')
        ip  = d.get('ip','')
        if ip and ip not in ips:
            set_config('ip_whitelist', (ips+','+ip).strip(','))
        return jsonify({'ok':True})
    active = get_config('ip_control_active','0') == '1'
    ips    = [i for i in get_config('ip_whitelist','').split(',') if i.strip()]
    return jsonify({'ok':True,'active':active,'ips':ips})


# ── Mot de passe personnel ────────────────────────────────────────────────────

@bp.route('/api/mon-password', methods=['POST'])
@login_required
def api_mon_password():
    d   = request.get_json() or {}
    pwd = d.get('password','')
    uid = session.get('user_id')
    if not pwd or not uid:
        return jsonify({'ok':False,'msg':'Mot de passe vide'})
    db_exec("UPDATE utilisateurs SET password_hash=? WHERE id=?",
            (hash_pwd(pwd), uid))
    return jsonify({'ok':True})


# ── Snapshot pour Bot ─────────────────────────────────────────────────────────

@bp.route('/api/param/snapshot/save', methods=['POST'])
@login_required
def api_snapshot_save():
    import json as _json
    d = request.get_json() or {}
    for k, v in d.items():
        db_exec(
            "INSERT OR REPLACE INTO data_snapshot(key,value,updated_at)"
            " VALUES(?,?,datetime('now'))",
            (k, _json.dumps(v, ensure_ascii=False, default=str)))
    return jsonify({'ok':True})


@bp.route('/api/param/snapshot/<key>')
@login_required
def api_snapshot_get(key):
    import json as _json
    row = db_one("SELECT value, updated_at FROM data_snapshot WHERE key=?", (key,))
    if not row:
        return jsonify({'ok':False,'msg':'Snapshot non trouve'})
    try:
        data = _json.loads(row['value'])
    except Exception:
        data = row['value']
    return jsonify({'ok':True,'data':data,'updated_at':row['updated_at']})


# ── Licence ───────────────────────────────────────────────────────────────────

@bp.route('/api/licence/statut')
@login_required
def api_licence_statut():
    from core.nexora_licence import get_licence_manager
    mgr  = get_licence_manager()
    if not mgr:
        return jsonify({'ok':False,'erreur':'Non initialise'})
    info = mgr.get_licence()
    if not info.valide and mgr.mode_demonstration():
        return jsonify({'ok':True,'mode':'DEMO','jours_demo':mgr.jours_demo_restants(),
                        'modules':list(PERMISSIONS_TREE.keys()),'nb_postes':1})
    if not info.valide:
        return jsonify({'ok':False,'mode':'INVALIDE','erreur':info.erreur})
    return jsonify({'ok':True,'mode':'PERPETUELLE' if info.perpetuelle else 'ACTIVE',
                    'nom_societe':info.nom_societe,'modules':info.modules,
                    'modules_noms':info.modules_noms,'nb_postes':info.nb_postes,
                    'date_expiration':str(info.date_expiration) if info.date_expiration else None,
                    'perpetuelle':info.perpetuelle,'jours_restants':info.jours_restants,
                    'expire_bientot':info.expire_bientot,'reference':info.reference})


@bp.route('/api/licence/activer', methods=['POST'])
@login_required
def api_licence_activer():
    from core.nexora_licence import get_licence_manager
    mgr    = get_licence_manager()
    d      = request.get_json() or {}
    numero = d.get('numero_serie','').strip()
    if not numero:
        return jsonify({'ok':False,'message':'Numero de serie manquant'})
    info = mgr.activer(numero)
    if info.valide:
        audit(session.get('username','?'), 'LICENCE', 'ACTIVATION', 'societe='+info.nom_societe)
        return jsonify({'ok':True,'message':'Licence activee pour '+info.nom_societe,
                        'nom_societe':info.nom_societe,'modules':info.modules})
    return jsonify({'ok':False,'message':info.erreur})


@bp.route('/api/licence/verifier', methods=['POST'])
def api_licence_verifier():
    from core.nexora_licence import NexoraLicenceVerifier
    d      = request.get_json() or {}
    numero = d.get('numero_serie','').strip()
    v      = NexoraLicenceVerifier()
    info   = v.verifier(numero)
    if info.valide:
        return jsonify({'ok':True,'nom_societe':info.nom_societe,
                        'modules_noms':info.modules_noms,'nb_postes':info.nb_postes,
                        'perpetuelle':info.perpetuelle,
                        'date_expiration':str(info.date_expiration) if info.date_expiration else None,
                        'jours_restants':info.jours_restants})
    return jsonify({'ok':False,'message':info.erreur})
